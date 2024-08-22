import requests
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Alignment


##################### SHORT URLS ############################

def shorten_url_tinyurl(long_url):
    url = f"http://tinyurl.com/api-create.php?url={long_url}"
    response = requests.get(url)
    if response.status_code == 200:
        return response.text
    else:
        print("Error al acortar la URL:", response.text)
        return long_url

##################### FORMAT functions ############################

# Formatear las columnas de listas como viñetas
def format_list_as_bullets(lst):
    try:
        if len(lst)==0: 
            return ""
        items = lst
        return "<ul>" + "".join(f"<li>{item}</li>" for item in items) + "</ul>"
    except Exception as e:
        return ""
    
# Formatear las columnas de listas como viñetas
def format_row_as_html(row):
    try:
        html_output = f"""<div class='producto'>
            <p><b>Descripción:</b></p>
                <p>{row['Descripcion']}</p>
            <p><b>Especificaciones:</b></p>
            <ul>
        """
        for especificacion in row['Especificaciones']:
            html_output += f"    <li>{especificacion}</li>\n"
        html_output += "  </ul>\n"
        
        html_output += "  <p><b>Ventajas:</b></p>\n"
        html_output += "  <ul>\n"
        for ventaja in row['Ventajas']:
            html_output += f"    <li>{ventaja}</li>\n"
        html_output += "  </ul>\n"
        html_output += "</div>\n"
        return html_output
    except Exception as e:
        return ""

# Formatear las columnas de listas como viñetas y enlaces clicables
def format_list_as_bullets_and_links(lst):
    try:
        if len(lst)==0:
            return ""
        items = lst 
        return "<ul>" + "".join(
        f"<li><a href='{item}' target='_blank'>{f'Enlace {i+1}'}</a></li>"
        for i, item in enumerate(items)
    ) + "</ul>"
    except Exception as e:
        return ""
    
def format_all_descriptions(lst):
    try:
        respuesta = ""
        
    except Exception as e:
        return ""
    
##################### CHECK functions ############################

# Función para verificar columnas requeridas
def check_required_columns(df, required_columns):
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        return False, missing_columns
    return True, []

# Función para verificar filas vacías
def check_empty_rows(df):
    empty_rows = df[df.isnull().all(axis=1)]
    return empty_rows

# Función para verificar si un enlace es válido
def es_enlace_valido(url):
    try:
        response = requests.head(url, allow_redirects=True)
        return response.status_code == 200
    except requests.RequestException:
        return False

# Función para verificar todos los enlaces en una fila
def verificar_fila_enlaces(row, key1, key2):
    try:
        for link in row[key1] + row[key2]:
            if  (not es_enlace_valido(link)):
                return False
        return True
    except Exception as e:
        print("Enlace no valido")
        print(e)
        return False

##################### CONVERT functions ############################

# Opción para descargar el archivo con descripciones
def convert_df_to_xlsx_1(dataframe):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        dataframe.to_excel(writer, index=False, sheet_name='Sheet1')
    return output.getvalue()

def calculate_row_height(text, font_size=11, line_height=15):
    # Cuenta los saltos de línea para estimar la cantidad de líneas
    line_count = text.count('\n') + 1  # +1 porque una línea sin \n sigue siendo una línea
    return line_count * line_height

def convert_df_to_xlsx(dataframe):
    # Paso 1: Exportar el DataFrame a un archivo Excel en un objeto BytesIO
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        dataframe.to_excel(writer, index=False, sheet_name='Sheet1')
    
    # Cargar el archivo Excel desde el objeto BytesIO
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    ws = wb.active  # Asumimos que trabajamos con la hoja activa
    
    # Paso 2: Ajustar el ancho de la columna y la altura de las filas
    column_letter = 'D'  # Asumimos que "DescripcionHTML" está en la columna B
    max_length = 100

    for cell in ws[column_letter]:
        if cell.value:
            # Calcular la longitud máxima de la columna
            max_length = max_length #max(max_length, len(str(cell.value)))
            # Calcular la altura de la fila automáticamente
            row_height = calculate_row_height(cell.value)
            ws.row_dimensions[cell.row].height = row_height
            # Alinear el texto en la parte superior de la celda
            cell.alignment = Alignment(vertical='top')

    for columna in ['A', 'B', 'C']:
        for cell in ws[columna]:
            if cell.value:
                cell.alignment = Alignment(vertical='top')
    
    # Ajustar el ancho de la columna
    ws.column_dimensions[column_letter].width = max_length

    # Guardar el archivo ajustado nuevamente en el objeto BytesIO
    adjusted_output = BytesIO()
    wb.save(adjusted_output)
    adjusted_output.seek(0)
    
    return adjusted_output.getvalue()


##################### PARALLEL functions ############################


# Función principal para paralelizar la tarea
def parallel_apply_formatting(df, columns, func, num_workers=4):
    df_split = np.array_split(df, num_workers)
    # Lista para guardar las futuras particiones
    futures = []

    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        for part in df_split:
            for column in columns:
                futures.append(executor.submit(apply_formatting, part, column,func))
                
        # Esperar y recolectar los resultados, preservando los índices originales
        results = [future.result() for future in futures]
    
    # Concatenar las particiones preservando los índices
    final_df = pd.concat(results).sort_index()
    
    return final_df

# Función para aplicar verificar_fila_enlaces a una partición del DataFrame
def verificar_enlaces_partition(partition, enlaces_col, imagenes_col):
    partition['enlaces_validos'] = partition.apply(lambda x: verificar_fila_enlaces(x, enlaces_col, imagenes_col), axis=1)
    return partition

# Función principal para paralelizar la verificación
def parallel_verificar_enlaces(df, enlaces_col, imagenes_col, num_workers=4):
    df_split = np.array_split(df, num_workers)
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        futures = [executor.submit(verificar_enlaces_partition, part, enlaces_col, imagenes_col) for part in df_split]
        results = [future.result() for future in futures]
    return pd.concat(results)


# Función para aplicar web scraping en paralelo
def aplicar_en_paralelo(df, func, max_workers=5):
    try:
        print(f"Entra a ejecutar en paralelo - {func}")
        print([row for index, row in df.iterrows()])
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            print("Starting parallel processing----------------------------------------")
            resultados = list(executor.map(func, [row for index, row in df.iterrows()]))
            print("Finished parallel processing----------------------------------------")
        return resultados
    except Exception as e:
        print(e)
        return None

# Función para aplicar la formateo a una columna
def apply_formatting(df, column, func):
    df[column] = df[column].apply(func)
    return df